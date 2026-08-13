#!/usr/bin/env python3
"""
İlave Gümrük Vergisi (İGV) - EK-1 parser.
Kaynak: Ticaret Bakanlığı, İthalatta İlave Gümrük Vergisi Uygulanmasına İlişkin Karar (3351), EK-1.

EK-1'de sütun başlıkları ülke grubu ismi değil sayı (1,2,3, GTS(4,5,6), 7) olarak geçiyor,
ama son sütun HER ZAMAN "Diğer Ülkeler (DÜ)" - bunu İthalat Rejimi Kararı'nın II sayılı
listesindeki paralel yapıdan (aynı kararın 6. maddesi bu tabloların ülke grubu tanımlarının
İthalat Rejimi Kararı ile aynı olduğunu belirtiyor) doğruladık. Çin bu sütuna girer.

Çıktı: gtip.db içine igv_diger_ulkeler tablosu (gtip12, igv_orani_pct).
"""
import glob
import os
import sqlite3
import sys
import types

# Bu ortamda Python 3.15 alpha + Pillow ABI uyuşmazlığı openpyxl'in PIL import'unu
# kırıyor (görsel desteğine ihtiyacımız yok) - sahte modül ile bypass ediyoruz.
fake_pil = types.ModuleType("PIL")
fake_image = types.ModuleType("PIL.Image")


class _Image:
    pass


fake_image.Image = _Image
fake_pil.Image = fake_image
sys.modules.setdefault("PIL", fake_pil)
sys.modules.setdefault("PIL.Image", fake_image)

import openpyxl  # noqa: E402

RAW_DIR = os.path.join(os.path.dirname(__file__), "..", "data", "raw", "igv-extracted")
DB_PATH = os.path.join(os.path.dirname(__file__), "..", "data", "processed", "gtip.db")


def norm_gtip(v):
    """Hücre değerini 12 haneli GTİP koduna normalize eder (nokta/boşluk temizler)."""
    s = str(v).strip().replace(".", "").replace(" ", "")
    if s.endswith(".0"):
        s = s[:-2]
    if s.isdigit():
        return s.zfill(12)
    return None


def find_ek1():
    for fp in glob.glob(os.path.join(RAW_DIR, "**", "*.xlsx"), recursive=True):
        if "EK-1" in os.path.basename(fp) or "EK 1" in os.path.basename(fp):
            return fp
    raise FileNotFoundError("EK-1.xlsx bulunamadı")


def main():
    fp = find_ek1()
    print(f"Okunuyor: {fp}")
    wb = openpyxl.load_workbook(fp, data_only=True)
    sh = wb[wb.sheetnames[0]]

    rows = []
    for r in range(1, sh.max_row + 1):
        gtip_raw = sh.cell(r, 1).value
        if gtip_raw is None:
            continue
        code = norm_gtip(gtip_raw)
        if not code:
            continue
        # Son dolu sütun = Diğer Ülkeler (DÜ). Sondan başlayıp ilk sayısal hücreyi al.
        last_val = None
        for c in range(sh.max_column, 2, -1):
            v = sh.cell(r, c).value
            if isinstance(v, (int, float)):
                last_val = v
                break
        if last_val is None:
            continue
        rows.append((code, float(last_val)))

    print(f"Bulunan GTİP+İGV satırı: {len(rows)}")

    conn = sqlite3.connect(DB_PATH)
    conn.execute("DROP TABLE IF EXISTS igv_diger_ulkeler")
    conn.execute("""
        CREATE TABLE igv_diger_ulkeler (
            gtip12 TEXT PRIMARY KEY,
            igv_orani_pct REAL
        )
    """)
    conn.executemany("INSERT OR REPLACE INTO igv_diger_ulkeler VALUES (?,?)", rows)
    conn.commit()

    n = conn.execute("SELECT COUNT(*) FROM igv_diger_ulkeler").fetchone()[0]
    nz = conn.execute("SELECT COUNT(*) FROM igv_diger_ulkeler WHERE igv_orani_pct > 0").fetchone()[0]
    print(f"Veritabanına yazılan: {n} (>%0 oranlı: {nz})")
    conn.close()


if __name__ == "__main__":
    main()
