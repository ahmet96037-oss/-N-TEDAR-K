#!/usr/bin/env python3
"""
Tarım ürünleri (fasıl 1-24 + 35/45/53) için Gümrük Vergisi düzeltmesi.
Bu listelerde "DÜ" (Diğer Ülkeler) başlıklı açık bir sütun var — sanayi listelerindeki
gibi konumu sabit değil (kaç ülke grubu varsa ona göre kayıyor), bu yüzden her sayfada
"DÜ" hücresini arayıp konumunu dinamik olarak buluyoruz.

Kapsam: I Sayılı Liste (19 sayfa), II Sayılı Liste (04-24. Fasıllar), III Sayılı Liste
(işlenmiş tarım), IV Sayılı Liste (balıkçılık). V/VI/VII Sayılı Liste kasıtlı olarak
DIŞARIDA — bunlar genel oran değil, özel/koşullu nihai-kullanım indirim programları.
"""
import os
import sys
import types

fake = types.ModuleType("PIL")
fake.Image = None
sys.modules["PIL"] = fake
sys.modules["PIL.Image"] = types.ModuleType("PIL.Image")
import openpyxl  # noqa: E402
from decimal import Decimal  # noqa: E402

from dotenv import load_dotenv  # noqa: E402
load_dotenv(os.path.join(os.path.dirname(__file__), "..", ".env"))
from src.db import db  # noqa: E402

FILES = [
    "data/raw/rejim-extracted/I sayìlì Liste.xlsx",
    "data/raw/rejim-extracted/II sayılı Liste (04-24.Fasıllar).xlsx",
    "data/raw/rejim-extracted/III sayìlì Liste.xlsx",
    "data/raw/rejim-extracted/IV sayılı Liste.xlsx",
]


def norm(g) -> str:
    """GTİP hücresi Excel'de SAYI olarak saklanmışsa (fasıl 01-09), baştaki sıfır
    integer dönüşümünde kayboluyor (0301... -> 301...) — bu durumda BAŞA sıfır
    tamamlanmalı, sona değil. Hücre zaten metin/noktalı formattaysa (ör.
    '8701.10.00.00.00') sona tamamlama doğru davranıştır (kısa/başlık seviyeli kodlar için)."""
    digits = "".join(ch for ch in str(g) if ch.isdigit())
    if isinstance(g, (int, float)) and not isinstance(g, bool):
        return digits.zfill(12)[:12]
    return digits.ljust(12, "0")[:12]


def find_du_col(ws, max_row=8):
    for r in range(1, max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and v.strip() == "DÜ":
                return c
    return None


def parse_file(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    for ws in wb.worksheets:
        du_col = find_du_col(ws)
        if not du_col:
            print(f"  [{ws.title}] DÜ sütunu bulunamadı, atlanıyor")
            continue
        for row in ws.iter_rows():
            cell0 = row[0].value
            if cell0 is None:
                continue
            digits = "".join(ch for ch in str(cell0) if ch.isdigit())
            if len(digits) < 10:
                continue
            du_val = ws.cell(row[0].row, du_col).value
            if isinstance(du_val, (int, float, Decimal)):
                yield norm(cell0), float(du_val)


def main():
    conn = db()
    CHUNK = 500
    total = 0
    for path in FILES:
        if not os.path.exists(path):
            print(f"BULUNAMADI: {path}")
            continue
        print(f"=== {path} ===")
        pairs = list(parse_file(path))
        cur = conn._conn.cursor()
        for i in range(0, len(pairs), CHUNK):
            batch = pairs[i:i + CHUNK]
            values_sql = ",".join("(%s::text,%s::numeric)" for _ in batch)
            flat = [v for pair in batch for v in pair]
            cur.execute(
                f"""UPDATE gtips SET base_duty_pct = v.pct, base_duty_source = 'ithalat_rejimi_liste_tarim_du'
                    FROM (VALUES {values_sql}) AS v(gtip12, pct)
                    WHERE gtips.gtip12 = v.gtip12""",
                flat,
            )
        conn._conn.commit()
        print(f"  {len(pairs)} satır işlendi")
        total += len(pairs)

    print(f"\nTOPLAM: {total}")
    n_gercek = conn.execute(
        "SELECT COUNT(*) AS n FROM gtips WHERE base_duty_source IN ('ithalat_rejimi_ek2_du','ithalat_rejimi_liste_tarim_du')"
    ).fetchone()["n"]
    n_tahmini = conn.execute(
        "SELECT COUNT(*) AS n FROM gtips WHERE base_duty_source = 'tgtc_vergi_haddi_tahmini'"
    ).fetchone()["n"]
    print(f"Gerçek kaynaklı (toplam): {n_gercek}")
    print(f"Hâlâ tahmini: {n_tahmini}")
    conn.close()


if __name__ == "__main__":
    main()
